#!/usr/bin/env python3
"""
gather_predictions.py — collect World Cup 2026 prediction Excel files into one JSON
payload and (optionally) inject it straight into the interactive HTML dashboard.

This replaces the heavy master-workbook pipeline. Scoring is NOT done here:
the dashboard computes all points live in the browser. This script only does
what Python is needed for:

  1. Extract each user's predictions (group scores, knockout scores, pens, Premios)
  2. Resolve their predicted bracket (group rankings, best-thirds, W/RU codes)
  3. Validate (missing scores, unresolved KO draws, missing Premios picks)
  4. Emit predictions.json — and/or splice it into the dashboard HTML in place

Usage:
  # Collect everything in a folder, write JSON + validation report
  python gather_predictions.py ./user_excels -o predictions.json

  # Add a single new participant later: just drop the file in the folder and re-run
  python gather_predictions.py ./user_excels --inject porra_mundial_2026.html

  # Mix folder + individual files
  python gather_predictions.py ./user_excels extra/Luis.xlsx --inject dashboard.html
"""

from __future__ import annotations

import argparse
import json
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

SHEET_WORLD_CUP = "World Cup"
SHEET_MATCHES = "Matches"
SHEET_GROUPS = "Groups"
SHEET_PREMIOS = "Premios"
SHEET_ASSIGN_THIRD = "AssignThird"

POINTS = {
    "exact": 4, "winner": 2, "cruce": 3,
    "advancement": {"R32": 2, "R16": 4, "QF": 6, "SF": 9, "Final": 13},
    "tournamentWinner": 18, "goal": 1, "award": 7,
}


# ----------------------------------------------------------------------------
# Small helpers
# ----------------------------------------------------------------------------

def clean(v: Any) -> Any:
    if v is None:
        return ""
    if isinstance(v, str):
        return v.strip()
    return v


def as_int(v: Any) -> int | None:
    if isinstance(v, bool):
        return None
    if isinstance(v, int):
        return v
    if isinstance(v, float) and v.is_integer():
        return int(v)
    if isinstance(v, str):
        v = v.strip()
        if v.isdigit() or (v.startswith("-") and v[1:].isdigit()):
            return int(v)
    return None


def cell_addr(row: int, col: int) -> str:
    return f"{get_column_letter(col)}{row}"


def round_name(no: int) -> str:
    if 1 <= no <= 72:
        return "Group"
    if 73 <= no <= 88:
        return "R32"
    if 89 <= no <= 96:
        return "R16"
    if 97 <= no <= 100:
        return "QF"
    if 101 <= no <= 102:
        return "SF"
    if no == 103:
        return "Third place"
    if no == 104:
        return "Final"
    return "Unknown"


def canon(name: str) -> str:
    """Accent/case-insensitive key for player-name comparison (validation only)."""
    s = unicodedata.normalize("NFKD", str(name))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().lower()


# ----------------------------------------------------------------------------
# Template geometry: where scores live on the "World Cup" sheet
# ----------------------------------------------------------------------------

def build_score_cell_map() -> list[dict[str, str]]:
    m: list[dict[str, str]] = []
    # Group stage: 12 groups across x 6 match-rows down
    for id_row in [11, 16, 21, 26, 31, 36]:
        for id_col in range(1, 35, 3):
            m.append({"id_cell": cell_addr(id_row, id_col),
                      "s1": cell_addr(id_row + 3, id_col + 1),
                      "s2": cell_addr(id_row + 3, id_col + 2),
                      "p1": "", "p2": ""})
    # Round of 32
    for c in range(1, 48, 3):
        m.append({"id_cell": cell_addr(48, c), "s1": cell_addr(51, c + 1),
                  "s2": cell_addr(51, c + 2), "p1": cell_addr(53, c + 1),
                  "p2": cell_addr(53, c + 2)})
    # Round of 16
    for c in range(1, 24, 3):
        m.append({"id_cell": cell_addr(58, c), "s1": cell_addr(61, c + 1),
                  "s2": cell_addr(61, c + 2), "p1": cell_addr(63, c + 1),
                  "p2": cell_addr(63, c + 2)})
    # Quarter-finals
    for c in [2, 8, 14, 20]:
        m.append({"id_cell": cell_addr(68, c), "s1": cell_addr(71, c + 1),
                  "s2": cell_addr(71, c + 3), "p1": cell_addr(73, c + 1),
                  "p2": cell_addr(73, c + 3)})
    # Semi-finals
    for c in [5, 17]:
        m.append({"id_cell": cell_addr(78, c), "s1": cell_addr(81, c + 1),
                  "s2": cell_addr(81, c + 3), "p1": cell_addr(83, c + 1),
                  "p2": cell_addr(83, c + 3)})
    # Third place + Final
    m.append({"id_cell": "K86", "s1": "L89", "s2": "N89", "p1": "L91", "p2": "N91"})
    m.append({"id_cell": "K96", "s1": "L99", "s2": "N99", "p1": "L101", "p2": "N101"})
    return m


# ----------------------------------------------------------------------------
# Workbook readers (data + formula views needed: codes live behind formulas)
# ----------------------------------------------------------------------------

def read_language_team_map(wb) -> dict[int, str]:
    out: dict[int, str] = {}
    if "Language" not in wb.sheetnames:
        return out
    for row in wb["Language"].iter_rows(min_row=5, min_col=4, max_col=10, values_only=True):
        no = as_int(row[0])
        name = next((clean(c) for c in row[1:] if clean(c)), "")
        if no is not None and name and no not in out:
            out[no] = str(name)
    return out


def read_group_code_to_team(wb) -> dict[str, str]:
    out: dict[str, str] = {}
    if SHEET_GROUPS not in wb.sheetnames:
        return out
    no_to_name = read_language_team_map(wb)
    for row in wb[SHEET_GROUPS].iter_rows(min_row=1, min_col=1, max_col=5, values_only=True):
        code = clean(row[1] if len(row) > 1 else "")
        team_no = as_int(row[2] if len(row) > 2 else None)
        team = clean(row[3] if len(row) > 3 else "") or (no_to_name.get(team_no, "") if team_no else "")
        if isinstance(code, str) and re.match(r"^[A-L][1-4]$", code) and team:
            out[code] = str(team)
    return out


def resolve_ref(wb, v: Any) -> Any:
    """Resolve simple cross-sheet refs like =AssignThird!$G$4."""
    if not isinstance(v, str) or not v.startswith("="):
        return v
    m = re.match(r"^=([^!]+)!\$?([A-Z]+)\$?(\d+)$", v)
    if not m:
        return v
    sheet = m.group(1).replace("'", "")
    if sheet not in wb.sheetnames:
        return v
    return wb[sheet][f"{m.group(2)}{m.group(3)}"].value


def read_match_catalog(wb_data, wb_formula) -> dict[int, dict[str, Any]]:
    ws = wb_data[SHEET_MATCHES]
    wsf = wb_formula[SHEET_MATCHES]
    code_to_team = read_group_code_to_team(wb_data)
    rows_d = list(ws.iter_rows(min_row=4, min_col=2, max_col=10, values_only=True))
    rows_f = list(wsf.iter_rows(min_row=4, min_col=2, max_col=10, values_only=True))
    catalog: dict[int, dict[str, Any]] = {}
    for i, row in enumerate(rows_d):
        no = as_int(row[0])
        if no is None:
            continue
        f = rows_f[i] if i < len(rows_f) else row
        c1 = clean(row[1]) or clean(resolve_ref(wb_formula, f[1]))
        c2 = clean(row[2]) or clean(resolve_ref(wb_formula, f[2]))
        if isinstance(c1, str) and c1.startswith("="):
            c1 = clean(resolve_ref(wb_formula, c1))
        if isinstance(c2, str) and c2.startswith("="):
            c2 = clean(resolve_ref(wb_formula, c2))
        t1, t2 = clean(row[7]), clean(row[8])
        if not t1 and isinstance(c1, str):
            t1 = code_to_team.get(c1, "")
        if not t2 and isinstance(c2, str):
            t2 = code_to_team.get(c2, "")
        catalog[no] = {"code_1": c1, "code_2": c2, "team_1": t1, "team_2": t2}
    return catalog


def read_third_assignment(wb_formula) -> dict[str, dict[str, str]]:
    if SHEET_ASSIGN_THIRD not in wb_formula.sheetnames:
        return {}
    rows = list(wb_formula[SHEET_ASSIGN_THIRD].iter_rows(min_row=4, min_col=3, max_col=11, values_only=True))
    if not rows:
        return {}
    codes = [clean(v) for v in rows[0][1:9]]
    table: dict[str, dict[str, str]] = {}
    for row in rows[4:]:
        combo = clean(row[0])
        if not combo or not isinstance(combo, str):
            continue
        table[str(combo)] = {str(c): str(g) for c, g in zip(codes, row[1:9])
                             if clean(c) and clean(g)}
    return table


def extract_score_inputs(wb_data) -> dict[int, dict[str, Any]]:
    ws = wb_data[SHEET_WORLD_CUP]
    out: dict[int, dict[str, Any]] = {}
    for fallback, m in enumerate(build_score_cell_map(), start=1):
        no = as_int(ws[m["id_cell"]].value) or fallback
        out[no] = {
            "s1": clean(ws[m["s1"]].value), "s2": clean(ws[m["s2"]].value),
            "p1": clean(ws[m["p1"]].value) if m["p1"] else "",
            "p2": clean(ws[m["p2"]].value) if m["p2"] else "",
        }
    return out


# ----------------------------------------------------------------------------
# Bracket resolution (group rankings -> seeds -> best thirds -> W/RU codes)
# ----------------------------------------------------------------------------

def decide_winner(t1, t2, s1, s2, p1="", p2=""):
    s1, s2, p1, p2 = as_int(s1), as_int(s2), as_int(p1), as_int(p2)
    if s1 is None or s2 is None:
        return ""
    if s1 > s2:
        return t1
    if s2 > s1:
        return t2
    if p1 is not None and p2 is not None:
        if p1 > p2:
            return t1
        if p2 > p1:
            return t2
    return "Draw"


def build_seed_resolver(catalog, scores, assignment_table):
    groups: dict[str, dict[str, dict[str, Any]]] = {}

    def ensure(g, t):
        groups.setdefault(g, {}).setdefault(t, {"team": t, "group": g, "pts": 0, "gf": 0, "ga": 0, "gd": 0})

    for no in range(1, 73):
        md = catalog.get(no, {})
        code = str(md.get("code_1", ""))
        if re.match(r"^[A-L][1-4]$", code):
            g = code[0]
            for t in (clean(md.get("team_1", "")), clean(md.get("team_2", ""))):
                if t:
                    ensure(g, t)
    for no in range(1, 73):
        md, sc = catalog.get(no, {}), scores.get(no, {})
        t1, t2 = clean(md.get("team_1", "")), clean(md.get("team_2", ""))
        code = str(md.get("code_1", ""))
        s1, s2 = as_int(sc.get("s1")), as_int(sc.get("s2"))
        if not t1 or not t2 or s1 is None or s2 is None or not re.match(r"^[A-L][1-4]$", code):
            continue
        g = code[0]
        ensure(g, t1)
        ensure(g, t2)
        a, b = groups[g][t1], groups[g][t2]
        a["gf"] += s1; a["ga"] += s2; b["gf"] += s2; b["ga"] += s1
        a["gd"] = a["gf"] - a["ga"]; b["gd"] = b["gf"] - b["ga"]
        if s1 > s2:
            a["pts"] += 3
        elif s2 > s1:
            b["pts"] += 3
        else:
            a["pts"] += 1; b["pts"] += 1

    key = lambda r: (-r["pts"], -r["gd"], -r["gf"], r["team"])
    seed: dict[str, str] = {}
    thirds: list[dict[str, Any]] = []
    for g in sorted(groups):
        ranked = sorted(groups[g].values(), key=key)
        if len(ranked) > 0:
            seed[f"1{g}"] = ranked[0]["team"]
        if len(ranked) > 1:
            seed[f"2{g}"] = ranked[1]["team"]
        if len(ranked) > 2:
            seed[f"3{g}"] = ranked[2]["team"]
            thirds.append(ranked[2])
    thirds.sort(key=key)
    combo = "".join(sorted(r["group"] for r in thirds[:8]))
    third_assignment = assignment_table.get(combo, {})

    def pick_third(slot: str, used: set[str]) -> str:
        ag = third_assignment.get(slot)
        if ag and f"3{ag}" in seed:
            used.add(ag)
            return seed[f"3{ag}"]
        for r in thirds:
            if r["group"] in slot.replace("3-", "") and r["group"] not in used:
                used.add(r["group"])
                return r["team"]
        return ""

    return seed, pick_third


# ----------------------------------------------------------------------------
# Per-user extraction + validation
# ----------------------------------------------------------------------------

def extract_user(path: Path, assignment_table: dict | None) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = []
    wb_data = load_workbook(path, data_only=True, read_only=True, keep_links=False)
    for sheet in (SHEET_WORLD_CUP, SHEET_MATCHES):
        if sheet not in wb_data.sheetnames:
            raise ValueError(f"missing sheet '{sheet}'")
    wb_formula = load_workbook(path, data_only=False, read_only=True, keep_links=False)

    catalog = read_match_catalog(wb_data, wb_formula)
    if assignment_table is None:
        assignment_table = read_third_assignment(wb_formula)
    scores = extract_score_inputs(wb_data)
    seed, pick_third = build_seed_resolver(catalog, scores, assignment_table)

    winners: dict[int, str] = {}
    losers: dict[int, str] = {}
    used_thirds: set[str] = set()

    def resolve(code: Any) -> str:
        code = str(clean(code))
        if not code:
            return ""
        if code.startswith("W") and code[1:].isdigit():
            return winners.get(int(code[1:]), "")
        if code.startswith("RU") and code[2:].isdigit():
            return losers.get(int(code[2:]), "")
        if code.startswith("3-"):
            return pick_third(code, used_thirds)
        if re.match(r"^[12][A-L]$", code):
            return seed.get(code, "")
        return ""

    matches: list[dict[str, Any]] = []
    missing_group, missing_ko, unresolved_ko, ko_draws = [], [], [], []
    for no in range(1, 105):
        md, sc = catalog.get(no, {}), scores.get(no, {})
        rnd = round_name(no)
        if no <= 72:
            t1, t2 = clean(md.get("team_1", "")), clean(md.get("team_2", ""))
        else:
            t1, t2 = resolve(md.get("code_1", "")), resolve(md.get("code_2", ""))
        w = decide_winner(t1, t2, sc.get("s1"), sc.get("s2"), sc.get("p1"), sc.get("p2"))
        loser = t2 if w == t1 else (t1 if w == t2 else "")
        if w and w != "Draw":
            winners[no] = w
        if loser:
            losers[no] = loser

        s1, s2 = as_int(sc.get("s1")), as_int(sc.get("s2"))
        if s1 is None or s2 is None:
            (missing_group if no <= 72 else missing_ko).append(no)
        elif no >= 73 and no != 103 and w == "Draw":
            ko_draws.append(no)
        if no >= 73 and (not t1 or not t2):
            unresolved_ko.append(no)

        matches.append({
            "match_no": no, "round": rnd, "team_1": t1, "team_2": t2,
            "score_1": "" if s1 is None else s1, "score_2": "" if s2 is None else s2,
            "penalty_1": as_int(sc.get("p1")) if as_int(sc.get("p1")) is not None else "",
            "penalty_2": as_int(sc.get("p2")) if as_int(sc.get("p2")) is not None else "",
            "winner": w,
        })

    bonus = {k: "" for k in ("top_scorer_1", "top_scorer_2", "top_scorer_3",
                             "top_scorer_4", "top_scorer_5", "golden_boot",
                             "golden_glove", "golden_ball")}
    if SHEET_PREMIOS in wb_data.sheetnames:
        ws = wb_data[SHEET_PREMIOS]
        for i, r in enumerate(range(5, 10), start=1):
            bonus[f"top_scorer_{i}"] = clean(ws[f"D{r}"].value)
        bonus["golden_boot"] = clean(ws["D10"].value)
        bonus["golden_glove"] = clean(ws["D11"].value)
        bonus["golden_ball"] = clean(ws["D12"].value)
    else:
        warnings.append("no Premios sheet — all award picks empty")
    bonus["tournament_winner"] = winners.get(104, "")

    if missing_group:
        warnings.append(f"group match(es) without a score: {_summarize(missing_group)}")
    if missing_ko:
        warnings.append(f"knockout match(es) without a score: {_summarize(missing_ko)}")
    if ko_draws:
        warnings.append(f"knockout draw(s) without penalty resolution at match(es) "
                        f"{_summarize(ko_draws)} — bracket cannot advance past them")
    if unresolved_ko:
        warnings.append(f"knockout slot(s) with unresolved team(s): "
                        f"{_summarize(unresolved_ko)}")
    empties = [k for k, v in bonus.items() if not clean(v)]
    if empties:
        warnings.append(f"empty Premios picks: {', '.join(empties)}")

    wb_data.close()
    wb_formula.close()
    return {"matches": matches, "bonus": bonus}, warnings


def _summarize(nos: list[int], limit: int = 12) -> str:
    s = ", ".join(map(str, nos[:limit]))
    return s + (f" … (+{len(nos) - limit} more)" if len(nos) > limit else "")


# ----------------------------------------------------------------------------
# Payload assembly + HTML injection
# ----------------------------------------------------------------------------

PAYLOAD_RE = re.compile(
    r'(<script[^>]*\bid="payload"[^>]*>)(.*?)(</script>)', re.DOTALL)

DATA_MARKER = "const DATA = "


def build_payload(user_data: dict[str, dict[str, Any]],
                  all_warnings: dict[str, list[str]]) -> dict[str, Any]:
    users = sorted(user_data)
    return {
        "points": POINTS,
        "users": users,
        "predictions": {u: user_data[u]["matches"] for u in users},
        "bonus": {u: user_data[u]["bonus"] for u in users},
        "warnings": {u: w for u, w in all_warnings.items() if w},
    }


def porra_preds(matches: list[dict[str, Any]]) -> dict[str, dict[str, Any]]:
    """Convert extracted matches to the dashboard's preds shape.

    Keyed by match number as string; only matches with both scores are
    included (same convention as the dashboard's embedded data)."""
    out: dict[str, dict[str, Any]] = {}
    for m in matches:
        if m["score_1"] == "" or m["score_2"] == "":
            continue
        out[str(m["match_no"])] = {
            "t1": m["team_1"], "t2": m["team_2"],
            "s1": m["score_1"], "s2": m["score_2"],
            "p1": m["penalty_1"] if m["penalty_1"] != "" else None,
            "p2": m["penalty_2"] if m["penalty_2"] != "" else None,
            "w": m["winner"],
        }
    return out


def porra_bonus(b: dict[str, Any]) -> dict[str, Any]:
    scorers = [clean(b.get(f"top_scorer_{i}", "")) for i in range(1, 6)]
    return {
        "scorers": [s for s in scorers if s],
        "boot": clean(b.get("golden_boot", "")) or None,
        "glove": clean(b.get("golden_glove", "")) or None,
        "ball": clean(b.get("golden_ball", "")) or None,
        "winner": clean(b.get("tournament_winner", "")) or None,
    }


def inject_into_porra(html_path: Path, user_data: dict[str, dict[str, Any]]) -> None:
    """Read-modify-write merge into the porra dashboard's `const DATA = {...};`.

    Replaces/adds preds + bonus per user, keeps schedule/teams/players/alias,
    and extends the players canonical-name map with unseen names so the
    dashboard's goal-entry table stays unified."""
    doc = html_path.read_text(encoding="utf-8")
    i = doc.index(DATA_MARKER) + len(DATA_MARKER)
    data, end = json.JSONDecoder().raw_decode(doc, i)

    for user, d in user_data.items():
        data.setdefault("preds", {})[user] = porra_preds(d["matches"])
        data.setdefault("bonus", {})[user] = porra_bonus(d["bonus"])
    data["users"] = sorted(set(data.get("users", [])) | set(user_data))

    players = data.setdefault("players", {})
    alias = data.setdefault("alias", {})
    known = set(players) | set(alias)
    display_to_key = {canon(v): k for k, v in players.items()}
    added: list[str] = []
    aliased: list[str] = []
    import difflib
    for d in user_data.values():
        pb = porra_bonus(d["bonus"])
        for name in pb["scorers"] + [pb["boot"], pb["glove"], pb["ball"]]:
            if not name:
                continue
            c = canon(name)
            if not c or c in known:
                continue
            # Same player written in full where a short key already exists
            # (e.g. "Mikel Oyarzabal" vs existing key "oyarzabal")
            if c in display_to_key:
                alias[c] = display_to_key[c]
                aliased.append(f"{name} -> {players[display_to_key[c]]}")
            else:
                # Likely typo of an existing player (e.g. "Erling Halland")
                close = difflib.get_close_matches(c, list(players), n=1, cutoff=0.92)
                if close:
                    alias[c] = close[0]
                    aliased.append(f"{name} -> {players[close[0]]} (fuzzy)")
                else:
                    players[c] = str(name)
                    display_to_key[c] = c
                    added.append(str(name))
            known.add(c)

    new_doc = doc[:i] + json.dumps(data, ensure_ascii=False) + doc[end:]
    backup = html_path.with_suffix(html_path.suffix + ".bak")
    backup.write_text(doc, encoding="utf-8")
    html_path.write_text(new_doc, encoding="utf-8")
    print(f"Merged {len(user_data)} user(s) into {html_path}  "
          f"({len(data['users'])} total players; backup: {backup.name})")
    if aliased:
        print(f"  Name variants auto-aliased to existing players: {'; '.join(sorted(aliased))}")
    if added:
        print(f"  New player names added to canonical map (check for typos / "
              f"add aliases if needed): {', '.join(sorted(added))}")


def inject_into_html(html_path: Path, user_data: dict[str, dict[str, Any]],
                     payload: dict[str, Any]) -> None:
    doc = html_path.read_text(encoding="utf-8")
    if DATA_MARKER in doc:
        inject_into_porra(html_path, user_data)
        return
    import html as html_mod
    blob = html_mod.escape(json.dumps(payload, ensure_ascii=False, separators=(",", ":")))
    new_doc, n = PAYLOAD_RE.subn(lambda m: m.group(1) + blob + m.group(3), doc)
    if n == 0:
        raise SystemExit(
            f'{html_path}: found neither `const DATA = ` nor a '
            f'<script id="payload"> block — is this the right dashboard file?')
    backup = html_path.with_suffix(html_path.suffix + ".bak")
    backup.write_text(doc, encoding="utf-8")
    html_path.write_text(new_doc, encoding="utf-8")
    print(f"Injected payload into {html_path}  (backup: {backup.name})")


# ----------------------------------------------------------------------------
# CLI
# ----------------------------------------------------------------------------

def gather_paths(inputs: list[Path]) -> list[Path]:
    files: list[Path] = []
    for p in inputs:
        if p.is_dir():
            files += sorted(q for q in p.glob("*.xls*")
                            if q.suffix.lower() in {".xlsx", ".xlsm"}
                            and not q.name.startswith("~$"))
        elif p.is_file():
            files.append(p)
        else:
            raise SystemExit(f"not found: {p}")
    if not files:
        raise SystemExit("no .xlsx/.xlsm prediction files found")
    # de-dup, keep order
    seen, out = set(), []
    for f in files:
        if f.resolve() not in seen:
            seen.add(f.resolve())
            out.append(f)
    return out


def main() -> None:
    ap = argparse.ArgumentParser(description="Collect prediction Excels into a dashboard payload.")
    ap.add_argument("inputs", nargs="+", type=Path,
                    help="Folder(s) and/or individual prediction .xlsx files")
    ap.add_argument("-o", "--output", type=Path, default=Path("predictions.json"),
                    help="JSON payload output (default: predictions.json)")
    ap.add_argument("--inject", type=Path, default=None,
                    help="Dashboard HTML to update in place (payload block is replaced; .bak kept)")
    ap.add_argument("--template", type=Path, default=None,
                    help="Optional template/actual workbook to read the AssignThird table from once")
    args = ap.parse_args()

    assignment_table = None
    if args.template:
        wbf = load_workbook(args.template, data_only=False, read_only=True, keep_links=False)
        assignment_table = read_third_assignment(wbf)
        wbf.close()

    files = gather_paths(args.inputs)
    user_data: dict[str, dict[str, Any]] = {}
    all_warnings: dict[str, list[str]] = {}
    failures: list[str] = []

    for f in files:
        user = f.stem
        try:
            data, warns = extract_user(f, assignment_table)
            user_data[user] = data
            all_warnings[user] = warns
            n_pred = sum(1 for m in data["matches"] if m["score_1"] != "" and m["score_2"] != "")
            status = "OK " if not warns else "WARN"
            print(f"[{status}] {user:<20} {n_pred:>3}/104 scored predictions, "
                  f"{len(warns)} warning(s)")
            for w in warns:
                print(f"        - {w}")
        except Exception as exc:
            failures.append(f"{f.name}: {exc}")
            print(f"[FAIL] {f.name}: {exc}", file=sys.stderr)

    if not user_data:
        raise SystemExit("nothing extracted — all files failed")

    payload = build_payload(user_data, all_warnings)
    if failures:
        payload["errors"] = failures

    args.output.write_text(json.dumps(payload, ensure_ascii=False, indent=1), encoding="utf-8")
    print(f"\nWrote {args.output}  ({args.output.stat().st_size/1024:.0f} KB, "
          f"{len(user_data)} user(s))")

    if args.inject:
        inject_into_html(args.inject, user_data, payload)


if __name__ == "__main__":
    main()
